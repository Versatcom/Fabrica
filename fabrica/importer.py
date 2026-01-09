from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from fabrica.templates import get_template
from fabrica.validation import ValidationResult, validate_rows


@dataclass(frozen=True)
class ImportReport:
    template_key: str
    total_rows: int
    result: ValidationResult


def import_from_csv(template_key: str, path: str | Path) -> ImportReport:
    template = get_template(template_key)
    rows = _read_csv(path)
    result = validate_rows(template, rows)
    return ImportReport(
        template_key=template.key,
        total_rows=len(rows),
        result=result,
    )


def import_from_excel(template_key: str, path: str | Path, sheet_name: str | None = None) -> ImportReport:
    template = get_template(template_key)
    rows = _read_excel(path, sheet_name)
    result = validate_rows(template, rows)
    return ImportReport(
        template_key=template.key,
        total_rows=len(rows),
        result=result,
    )


def _read_csv(path: str | Path) -> List[Dict[str, str]]:
    resolved = Path(path)
    with resolved.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_excel(path: str | Path, sheet_name: str | None) -> List[Dict[str, str]]:
    resolved = Path(path)
    workbook = load_workbook(resolved, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: List[Dict[str, str]] = []
    for row in rows[1:]:
        record = {
            header: "" if value is None else str(value).strip()
            for header, value in zip(headers, row)
        }
        records.append(record)
    return records
